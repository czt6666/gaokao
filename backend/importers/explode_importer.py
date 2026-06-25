"""
新版 xlsx Explode 导入器

把每行多年 xlsx 数据 "explode" 成 admission_records（一年一行），
同时写入 admission_2026（2026 计划数据，一行多年格式）。

用法（在 backend 目录）:
  # Dry-run 扫描
  .venv/bin/python -m importers.explode_importer --dir "/Users/czt/workspace/webfrontend/高考程序素材/2026"

  # 正式导入
  .venv/bin/python -m importers.explode_importer --dir "<目录>" --apply

核心流程:
  1. ColumnMap.detect() 识别每省 xlsx 列对应关系
  2. 逐行读取，应用静态映射表（static_maps）
  3. Explode: 一行 → 至多 4 行（2026/2025/2024/2023）
  4. 写 admission_2026（2026计划） + admission_records（历年数据）
"""
import argparse
import glob
import hashlib
import json
import os
import sqlite3
import time

import openpyxl

from .field_registry import PLAN_FIELDS
from .column_mapper import ColumnMap
from .static_maps import (
    get_batch_type,
    get_subject_must,
    get_subject_any_of,
    get_major_restrictions,
    get_is_985_211,
    get_derived_category,
    BATCH_TYPE_MAP,
    SUBJECT_MUST_MAP,
    SUBJECT_ANY_OF_MAP,
    PLAN_CATEGORY_MAP,
)

DEFAULT_DB = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "gaokao.db")

# ── 工具函数 ──────────────────────────────────────────────────────────────
def _s(v):
    return "" if v is None else str(v).strip()


def _i(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return int(v)
    t = str(v).strip().replace(",", "")
    if t in ("", "-", "—", "待定", "/", "无", "N/A", "暂无"):
        return None
    try:
        return int(float(t))
    except ValueError:
        return None


def _cell(row, idx):
    return row[idx] if idx is not None and idx < len(row) else None


# ── admission_records DDL（新增列）──────────────────────────────────────
ADMISSION_NEW_COLS = [
    ("major_full",        "VARCHAR(200) DEFAULT ''"),
    ("major_code",        "VARCHAR(50) DEFAULT ''"),
    ("group_name",        "VARCHAR(100) DEFAULT ''"),
    ("group_code",        "VARCHAR(50) DEFAULT ''"),
    ("group_full_code",   "VARCHAR(50) DEFAULT ''"),
    ("discipline",        "VARCHAR(100) DEFAULT ''"),
    ("major_class",       "VARCHAR(100) DEFAULT ''"),
    ("major_level",       "VARCHAR(20) DEFAULT ''"),
    ("plan_count",        "INTEGER"),
    ("plan_category",     "VARCHAR(100) DEFAULT ''"),
    ("est_rank_26",       "INTEGER"),
    ("school_tags",       "VARCHAR(200) DEFAULT ''"),
    ("school_level",      "VARCHAR(100) DEFAULT ''"),
    ("city",              "VARCHAR(50) DEFAULT ''"),
    ("city_level",        "VARCHAR(50) DEFAULT ''"),
    ("duration",          "VARCHAR(20) DEFAULT ''"),
    ("tuition",           "VARCHAR(50) DEFAULT ''"),
    ("ruanke_grade",      "VARCHAR(20) DEFAULT ''"),
    ("ruanke_rank",       "VARCHAR(20) DEFAULT ''"),
    ("discipline_eval",   "VARCHAR(20) DEFAULT ''"),
    ("major_level_tag",   "VARCHAR(50) DEFAULT ''"),
    ("source_file",       "VARCHAR(200) DEFAULT ''"),
]

ADMISSION_EXISTING_COLS = [
    "school_code", "school_name", "major_name", "major_group",
    "province", "year", "batch", "subject_req", "min_score", "min_rank",
    "admit_count", "school_province", "school_nature", "is_985", "is_211",
    "batch_type", "subject_must", "subject_any_of", "major_restrictions",
    "major_remark", "derived_category",
]

# admission_2026 extra cols
A26_EXTRA_COLS = [
    ("source_file", "TEXT"),
    ("row_uid", "TEXT"),
]


def ensure_schema(conn: sqlite3.Connection):
    """确保 admission_records 和 admission_2026 所需的列存在。"""
    cur = conn.cursor()

    # admission_records
    existing = {r[1] for r in cur.execute("PRAGMA table_info(admission_records)")}
    for col_name, col_def in ADMISSION_NEW_COLS:
        if col_name not in existing:
            print(f"  [ALTER] admission_records ADD {col_name}")
            cur.execute(f"ALTER TABLE admission_records ADD COLUMN {col_name} {col_def}")
    conn.commit()

    # admission_2026
    cur.execute("""
        CREATE TABLE IF NOT EXISTS admission_2026 (
            id INTEGER PRIMARY KEY AUTOINCREMENT
        )
    """)
    existing_a26 = {r[1] for r in cur.execute("PRAGMA table_info(admission_2026)")}
    for f in PLAN_FIELDS:
        if f.key not in existing_a26:
            cur.execute(f"ALTER TABLE admission_2026 ADD COLUMN {f.key} {'INTEGER' if f.type == 'int' else 'TEXT'}")
    for name, typ in A26_EXTRA_COLS:
        if name not in existing_a26:
            cur.execute(f"ALTER TABLE admission_2026 ADD COLUMN {name} {typ}")
    conn.commit()


def explode_file(path: str, conn: sqlite3.Connection, apply: bool, stats: dict):
    """处理单个 xlsx 文件：detect → explode → 写入。"""
    cm = ColumnMap.detect(path)
    prov = cm.province
    if cm.warnings:
        for w in cm.warnings:
            print(f"  ⚠ [{prov}] {w}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    fc = cm.field_cols
    hist_blocks = cm.history_blocks()
    src = os.path.basename(path)

    a26_rows = []
    adm_rows = []
    seen_uid = set()
    seen_hist = set()

    for row in ws.iter_rows(min_row=cm.data_start_row, values_only=True):
        sch = _s(_cell(row, fc.get("school_name")))
        if not sch:
            continue

        # ── 基础字段提取 ──────────────────────────────────────────────
        school_code = _s(_cell(row, fc.get("school_code")))
        school_name = sch
        major_name = _s(_cell(row, fc.get("major_name")))
        major_full = _s(_cell(row, fc.get("major_full")))
        major_code = _s(_cell(row, fc.get("major_code")))
        major_remark = _s(_cell(row, fc.get("major_remark")))
        province = _s(_cell(row, fc.get("province"))) or prov
        batch = _s(_cell(row, fc.get("batch")))
        category = _s(_cell(row, fc.get("category")))
        subject_req = _s(_cell(row, fc.get("subject_req")))
        major_level = _s(_cell(row, fc.get("major_level")))
        plan_count = _i(_cell(row, fc.get("plan_count")))
        duration = _s(_cell(row, fc.get("duration")))
        tuition = _s(_cell(row, fc.get("tuition")))
        discipline = _s(_cell(row, fc.get("discipline")))
        major_class_s = _s(_cell(row, fc.get("major_class")))
        est_rank_26 = _i(_cell(row, fc.get("est_rank_26")))
        is_new = _s(_cell(row, fc.get("is_new")))

        # 院校信息
        school_province = _s(_cell(row, fc.get("school_province")))
        city = _s(_cell(row, fc.get("city")))
        city_level = _s(_cell(row, fc.get("city_level")))
        school_tags = _s(_cell(row, fc.get("school_tags")))
        school_level = _s(_cell(row, fc.get("school_level")))
        nature = _s(_cell(row, fc.get("nature")))
        group_name = _s(_cell(row, fc.get("group_name")))
        group_code = _s(_cell(row, fc.get("group_code")))
        group_full_code = _s(_cell(row, fc.get("group_full_code")))
        plan_category = _s(_cell(row, fc.get("plan_category")))
        ruanke_grade = _s(_cell(row, fc.get("ruanke_grade")))
        ruanke_rank = _s(_cell(row, fc.get("ruanke_rank")))
        discipline_eval = _s(_cell(row, fc.get("discipline_eval")))
        major_level_tag = _s(_cell(row, fc.get("major_level_tag")))

        # ── 静态映射 ──────────────────────────────────────────────────
        batch_type = get_batch_type(batch)
        is_985, is_211 = get_is_985_211(school_tags)
        derived_category = get_derived_category(category)
        subject_must = get_subject_must(subject_req)

        # 3+1+2 省份不限选科 → must 补首选科目
        if (not subject_must) and category in ("物理", "历史", "物理类", "历史类", "理科", "文科"):
            first_map = {"物理": "物理", "物理类": "物理", "理科": "物理",
                         "历史": "历史", "历史类": "历史", "文科": "历史"}
            subject_must = first_map.get(category, "")

        subject_any_of = get_subject_any_of(subject_req)
        restrictions_tags = get_major_restrictions(plan_category, major_full, major_remark)
        major_restrictions = ",".join(restrictions_tags) if restrictions_tags else ""

        # ── admission_2026 行 ──────────────────────────────────────
        # uid: province + school_code + group_full_code + major_code
        uid_basis = "|".join([province, school_code, group_full_code, major_code])
        row_uid = hashlib.md5(uid_basis.encode()).hexdigest()

        if row_uid not in seen_uid:
            seen_uid.add(row_uid)
            a26_rec = {}
            for f in PLAN_FIELDS:
                v = _cell(row, fc.get(f.key))
                a26_rec[f.key] = _i(v) if f.type == "int" else _s(v)
            a26_rec["province"] = a26_rec.get("province") or province
            a26_rec["source_file"] = src
            a26_rec["row_uid"] = row_uid
            a26_rows.append(a26_rec)

        # ── Explode 历年数据 → admission_records ───────────────────
        plan_year = _i(_cell(row, fc.get("plan_year"))) or 2026

        # 共享的元数据（所有年份行共用）
        shared = {
            "school_code": school_code,
            "school_name": school_name,
            "major_name": major_name,
            "major_full": major_full,
            "major_code": major_code,
            "major_group": group_name,
            "major_remark": major_remark,
            "province": province,
            "batch": batch,
            "subject_req": subject_req,
            "batch_type": batch_type,
            "subject_must": subject_must,
            "subject_any_of": subject_any_of,
            "major_restrictions": major_restrictions,
            "derived_category": derived_category,
            "school_province": school_province,
            "school_nature": nature,
            "is_985": is_985,
            "is_211": is_211,
            "school_tags": school_tags,
            "school_level": school_level,
            "city": city,
            "city_level": city_level,
            "major_level": major_level,
            "discipline": discipline,
            "major_class": major_class_s,
            "group_name": group_name,
            "group_code": group_code,
            "group_full_code": group_full_code,
            "plan_category": plan_category,
            "duration": duration,
            "tuition": tuition,
            "ruanke_grade": ruanke_grade,
            "ruanke_rank": ruanke_rank,
            "discipline_eval": discipline_eval,
            "major_level_tag": major_level_tag,
            "source_file": src,
        }

        # 注意：2026 预估位次（est_rank_26）只保留在 admission_2026 表，
        # 不写入 admission_records。理由：
        #   ① 2026 是预测值，不是真实录取，混进历年录取位次表会误导用户；
        #   ② 推荐的概率预测应只基于真实历年(2023-2025)数据；
        #   ③ 预测排名在前端学校页单独高亮展示（取自 admission_2026.est_rank_26）。

        # 历年数据（仅真实录取年份 2023-2025）
        for idx, year, cols in hist_blocks:
            min_score = _i(_cell(row, cols.get("min_score")))
            if min_score is None:
                continue
            yr = cm.row_year(idx, row)
            admit_count = _i(_cell(row, cols.get("admit_count")))
            min_rank = _i(_cell(row, cols.get("min_rank"))) or 0

            hist_key = (province, school_name, major_name, yr, min_score)
            if hist_key in seen_hist:
                continue
            seen_hist.add(hist_key)
            adm_rows.append({
                **shared,
                "year": yr,
                "min_score": min_score,
                "min_rank": min_rank,
                "admit_count": admit_count or 0,
                "plan_count": None,
                "est_rank_26": None,
            })

    wb.close()

    plan_cnt = len(a26_rows)
    hist_cnt = len(adm_rows)
    print(f"  [{prov}] plan={plan_cnt}  explode={hist_cnt}")
    stats["plan"] += plan_cnt
    stats["hist"] += hist_cnt

    if apply and (plan_cnt or hist_cnt):
        cur = conn.cursor()

        # 写 admission_2026
        if a26_rows:
            a26_cols = [f.key for f in PLAN_FIELDS] + [c[0] for c in A26_EXTRA_COLS]
            ph = ",".join("?" for _ in a26_cols)
            cur.executemany(
                f"INSERT OR REPLACE INTO admission_2026 ({','.join(a26_cols)}) VALUES ({ph})",
                [tuple(r.get(c) for c in a26_cols) for r in a26_rows])

        # 写 admission_records
        if adm_rows:
            all_cols = ADMISSION_EXISTING_COLS + [c[0] for c in ADMISSION_NEW_COLS
                                                   if c[0] not in ADMISSION_EXISTING_COLS]
            # 过滤掉不存在的列
            existing_cols = {r[1] for r in cur.execute("PRAGMA table_info(admission_records)")}
            write_cols = [c for c in all_cols if c in existing_cols]
            ph = ",".join("?" for _ in write_cols)
            cur.executemany(
                f"INSERT INTO admission_records ({','.join(write_cols)}) VALUES ({ph})",
                [tuple(r.get(c) for c in write_cols) for r in adm_rows])

        conn.commit()


# ── 主流程 ──────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser(description="新版xlsx Explode导入器")
    ap.add_argument("--dir", help="包含多省 xlsx 的目录")
    ap.add_argument("--file", help="单个 xlsx")
    ap.add_argument("--db", default=DEFAULT_DB)
    ap.add_argument("--apply", action="store_true", help="真正写库（默认 dry-run）")
    ap.add_argument("--check-batches", action="store_true", help="检查 batch_type 覆盖率后退出")
    args = ap.parse_args()

    files = []
    if args.file:
        files = [args.file]
    elif args.dir:
        files = sorted(glob.glob(os.path.join(args.dir, "*.xlsx")))
    if not files:
        raise SystemExit("需要 --dir 或 --file")

    # ── 批次映射覆盖率检查 ──────────────────────────────────
    if args.check_batches:
        all_b = set()
        for f in files:
            cm = ColumnMap.detect(f)
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            ws = wb.worksheets[0]
            fc = cm.field_cols
            batch_col = fc.get("batch")
            if batch_col is None:
                wb.close()
                continue
            for row in ws.iter_rows(min_row=cm.data_start_row, values_only=True):
                if batch_col < len(row) and row[batch_col]:
                    all_b.add(str(row[batch_col]).strip())
            wb.close()
        missing = [b for b in sorted(all_b) if b not in BATCH_TYPE_MAP]
        if missing:
            print(f"⚠ 未映射的 batch 值 ({len(missing)}):")
            for b in missing:
                print(f"    '{b}'")
        else:
            print(f"✓ 所有 {len(all_b)} 个 batch 值均已映射")
        return

    mode = "APPLY" if args.apply else "DRY-RUN"
    print(f"== {mode} | DB: {args.db} | 文件 {len(files)} 个 ==")

    conn = sqlite3.connect(args.db)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")

    if args.apply:
        ensure_schema(conn)
        print("  Schema 就绪")

        # 清空旧数据
        print("  清空旧数据...")
        conn.execute("DELETE FROM admission_records")
        conn.execute("DELETE FROM admission_2026")
        conn.execute("DELETE FROM sqlite_sequence WHERE name IN ('admission_records','admission_2026')")
        conn.commit()
        print("  旧数据已清空")

    stats = {"plan": 0, "hist": 0}
    t0 = time.time()
    for f in files:
        explode_file(f, conn, args.apply, stats)

    elapsed = time.time() - t0
    print(f"\n合计: plan {stats['plan']:,} 行, explode {stats['hist']:,} 行 ({elapsed:.0f}s)")

    if args.apply:
        # 创建索引
        print("创建索引...")
        cur = conn.cursor()
        cur.execute("CREATE INDEX IF NOT EXISTS ix_ar_prov_year ON admission_records(province, year)")
        cur.execute("CREATE INDEX IF NOT EXISTS ix_ar_school ON admission_records(school_name)")
        cur.execute("CREATE INDEX IF NOT EXISTS ix_ar_major ON admission_records(major_name)")
        cur.execute("CREATE UNIQUE INDEX IF NOT EXISTS ux_a26_uid ON admission_2026(row_uid)")
        cur.execute("CREATE INDEX IF NOT EXISTS ix_a26_prov ON admission_2026(province)")
        cur.execute("CREATE INDEX IF NOT EXISTS ix_a26_school ON admission_2026(school_name)")
        conn.commit()
        print("  索引就绪")
        print("✓ 导入完成")
    else:
        print("(dry-run 结束，加 --apply 执行)")

    conn.close()


if __name__ == "__main__":
    main()
