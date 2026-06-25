"""列号 ↔ 规范字段 对应层。

对单个 xlsx：
  1. 自动识别表头行（前 4 行里命中锚点字段最多的一行）；
  2. 把每个列号映射到规范字段（PLAN_FIELDS）；
  3. 识别历年录取块（最低分1/2/3，或重名裸「最低分」按位置兜底）；
  4. 未识别的表头进 extra（导入时落到 extra_json，绝不丢数据）。

输出 ColumnMap 对象，既供导入器使用，也可 describe() 打印人工核对。
"""
import re
import openpyxl
from openpyxl.utils import get_column_letter

from .field_registry import (
    ALIAS_TO_FIELD, HISTORY_BASENAMES, HISTBASE_TO_KEY, ANCHOR_HEADERS,
)

PLAN_YEAR = 2026  # 块序号 -> 年份： year = PLAN_YEAR - idx （块1=2025…）


class ColumnMap:
    def __init__(self, path):
        self.path = path
        self.province = ""
        self.header_row = 0          # 1-based
        self.data_start_row = 0
        self.field_cols = {}         # 规范字段key -> 0-based 列号
        self.blocks = {}             # 块序号(int) -> {admission_records列名: 0-based 列号}
        self.year_cols = {}          # 块序号 -> 年份列 0-based（仅个别省份有显式年份列）
        self.extra_cols = {}         # 未识别表头文本 -> 0-based 列号
        self.headers = []            # 表头行原始文本
        self.warnings = []           # 需人工复核的提示

    # ── 解析 ──────────────────────────────────────────────────────────────
    @classmethod
    def detect(cls, path):
        cm = cls(path)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        first4 = [list(r) for r in ws.iter_rows(min_row=1, max_row=4, values_only=True)]

        # 1) 表头行 = 锚点命中最多的一行
        def anchor_hits(row):
            return sum(1 for v in row if isinstance(v, str) and v.strip() in ANCHOR_HEADERS)
        hdr_i = max(range(len(first4)), key=lambda i: anchor_hits(first4[i]))
        cm.header_row = hdr_i + 1
        cm.data_start_row = hdr_i + 2
        cm.headers = [(str(v).strip() if v is not None else "") for v in first4[hdr_i]]

        # 2/3) 逐列分类
        # 历年块两种排布：①带序号「最低分1/2/3」→按序号分块；②裸重名「最低分/最低分/最低分」
        # （吉林/宁夏/河北）→位置聚类：从左到右扫，遇到“已在当前块出现过的基名”即开新块。
        cluster_idx = 0
        cluster_bases = set()
        bare_repeat = False
        for col, h in enumerate(cm.headers):
            if not h:
                continue
            digit = re.search(r"(\d+)$", h)
            base = re.sub(r"\s*\d+$", "", h).strip()

            if base in HISTORY_BASENAMES:               # 历年块列
                if digit:
                    idx = int(digit.group(1))
                else:                                    # 裸重名 → 位置聚类
                    bare_repeat = True
                    if base in cluster_bases:
                        cluster_idx += 1
                        cluster_bases = set()
                    if cluster_idx == 0:
                        cluster_idx = 1
                    idx = cluster_idx
                    cluster_bases.add(base)
                cm.blocks.setdefault(idx, {})[HISTBASE_TO_KEY[base]] = col
            elif base == "年份" and digit:              # 显式年份列（如北京 年份1/2/3）
                cm.year_cols[int(digit.group(1))] = col
            elif h in ALIAS_TO_FIELD:                   # 规范字段
                # 多次出现取首列：身份/2026计划信息（含「计划人数」）恒在历年块左侧，
                # 末列的「计划人数」属历年块，会误取（吉林/宁夏/云南/广西/江苏）。
                cm.field_cols.setdefault(ALIAS_TO_FIELD[h].key, col)
            else:                                       # 兜底
                cm.extra_cols[h] = col

        # 「计划人数」逐年重复出现时，plan_count 已取首列（2026计划区）。仅当首列仍落在
        # 历年块内（即计划区没有独立「计划人数」）才告警需人工复核。
        _hist_cols = [c for b in cm.blocks.values() for c in b.values()]
        _first_hist = min(_hist_cols) if _hist_cols else None
        _pc = cm.field_cols.get("plan_count")
        _pc_in_hist = _pc is None or (_first_hist is not None and _pc >= _first_hist)
        if bare_repeat and cm.headers.count("计划人数") > 1 and _pc_in_hist:
            cm.warnings.append(
                "历年块为裸重名排布，且「计划人数」多次出现，计划区未找到独立计划人数列："
                "plan_count 可能误取，建议为该省提供手工列映射覆盖（历年分数已按位置聚类，回填不受影响）。"
            )

        # 4) 省份：取首条数据行的「生源地」；缺失则用文件名兜底
        prov_col = cm.field_cols.get("province")
        if prov_col is not None:
            for r in ws.iter_rows(min_row=cm.data_start_row, max_row=cm.data_start_row + 5,
                                  values_only=True):
                if r and prov_col < len(r) and r[prov_col]:
                    cm.province = str(r[prov_col]).strip()
                    break
        if not cm.province:
            cm.province = _province_from_filename(path)
        wb.close()
        return cm

    # ── 历年块按年份排序，附带年份 ─────────────────────────────────────────
    def history_blocks(self):
        """返回 [(idx, year, {col_key: col_idx}), …]，仅含有 min_score 的块。"""
        out = []
        for idx in sorted(self.blocks):
            cols = self.blocks[idx]
            if "min_score" not in cols:
                continue
            out.append((idx, PLAN_YEAR - idx, cols))
        return out

    def row_year(self, idx, row):
        """块的年份：优先读显式年份列，否则按序号推算。"""
        yc = self.year_cols.get(idx)
        if yc is not None and yc < len(row) and row[yc]:
            try:
                return int(row[yc])
            except (TypeError, ValueError):
                pass
        return PLAN_YEAR - idx

    # ── 人工核对用 ─────────────────────────────────────────────────────────
    def describe(self):
        L = get_column_letter
        lines = [f"文件: {self.path.split('/')[-1]}",
                 f"省份: {self.province}  表头行: 第{self.header_row}行  数据起始: 第{self.data_start_row}行",
                 f"识别规范字段: {len(self.field_cols)}  历年块: {sorted(self.blocks)}  未识别列: {len(self.extra_cols)}",
                 "── 规范字段 列号对应 ──"]
        for k, c in sorted(self.field_cols.items(), key=lambda x: x[1]):
            lines.append(f"  {L(c+1):>3}({c}) -> {k}  [{self.headers[c]}]")
        lines.append("── 历年块 列号对应 ──")
        for idx, year, cols in self.history_blocks():
            cc = ", ".join(f"{k}={L(v+1)}" for k, v in cols.items())
            lines.append(f"  块{idx}=>{year}年: {cc}")
        if self.extra_cols:
            lines.append("── 未识别(进 extra_json) ──")
            lines.append("  " + ", ".join(f"{L(c+1)}={h}" for h, c in self.extra_cols.items()))
        return "\n".join(lines)


_PROV_KEYS = ["内蒙古", "黑龙江", "北京", "天津", "上海", "重庆", "河北", "山西", "辽宁",
              "吉林", "江苏", "浙江", "安徽", "福建", "江西", "山东", "河南", "湖北",
              "湖南", "广东", "广西", "海南", "四川", "贵州", "云南", "陕西", "甘肃",
              "青海", "宁夏", "新疆", "西藏"]


def _province_from_filename(path):
    base = path.split("/")[-1]
    for p in _PROV_KEYS:
        if p in base:
            return p
    return ""
