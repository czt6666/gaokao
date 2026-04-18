#!/usr/bin/env python3
"""
批量导入全国各省高考录取数据（专业分数线）到 gaokao.db
支持两种格式：
  格式A: "22-25年全国高校在{省}的专业录取分数.xlsx"（统一格式，优先）
  格式B: "XX_专业分数线_YYYY.xlsx"（旧格式，按年份）
"""
import os, sqlite3, time, glob
import openpyxl

DB_PATH = os.path.join(os.path.dirname(__file__), "gaokao.db")
DATA_ROOT = "/Users/Admin/Desktop/05、【琢玉攻略】高考数据"

# ── 格式A：统一22-25文件，字段映射 ────────────────────────────────
# 表头: 年份, 院校名称, 院校代码, 科类, 批次, 选科要求, 专业, 专业代码, 所属专业组, 专业备注, 录取人数, 最低分数, 最低位次, 学校所在, 学校性质, 是否985, 是否211
FORMAT_A_MAP = {
    "年份": "year", "院校名称": "school_name", "院校代码": "school_code",
    "科类": "batch", "批次": "batch2", "选科要求": "subject_req",
    "专业": "major_name", "所属专业组": "major_group",
    "录取人数": "admit_count", "最低分数": "min_score", "最低位次": "min_rank",
    "学校所在": "school_province", "学校性质": "school_nature",
    "是否985": "is_985", "是否211": "is_211",
}

# ── 格式B：旧版按年文件 ────────────────────────────────────────────
# 表头样例: 年份, 学校, 招生代码, 学校方向, 省份, 科目, 专业, 专业代码, 批次, ..., 最低分, 最低分位次, ..., 录取人数, 选科要求
FORMAT_B_MAP = {
    "年份": "year", "学校": "school_name", "招生代码": "school_code",
    "省份": "school_province", "科目": "batch", "批次": "batch2",
    "专业": "major_name", "最低分": "min_score", "最低分位次": "min_rank",
    "录取人数": "admit_count", "选科要求": "subject_req",
}

def safe_int(v):
    try:
        if v is None or v == '-' or v == '': return None
        return int(float(str(v).replace(',', '')))
    except: return None

def safe_str(v):
    if v is None: return ''
    return str(v).strip()


def import_format_a(ws, province: str, cur, existing: set) -> int:
    """导入格式A（统一22-25文件）"""
    rows = ws.iter_rows(values_only=True)
    header = [safe_str(h) for h in next(rows)]
    col = {h: i for i, h in enumerate(header)}

    def get(row, name):
        idx = col.get(name)
        return row[idx] if idx is not None else None

    inserted = 0
    batch = []
    for row in rows:
        year = safe_int(get(row, '年份'))
        school = safe_str(get(row, '院校名称'))
        major = safe_str(get(row, '专业'))
        min_rank = safe_int(get(row, '最低位次'))

        if not year or not school or not major or not min_rank:
            continue

        key = (province, year, school, major)
        if key in existing:
            continue
        existing.add(key)

        batch.append((
            safe_str(get(row, '院校代码')),
            school,
            major,
            safe_str(get(row, '所属专业组')),
            province,
            year,
            safe_str(get(row, '科类')) or safe_str(get(row, '批次')),
            safe_str(get(row, '选科要求')),
            safe_int(get(row, '最低分数')),
            min_rank,
            safe_int(get(row, '录取人数')),
            safe_str(get(row, '学校所在')),
            safe_str(get(row, '学校性质')),
            safe_str(get(row, '是否985')),
            safe_str(get(row, '是否211')),
        ))

        if len(batch) >= 2000:
            cur.executemany(INSERT_SQL, batch)
            inserted += len(batch)
            batch = []

    if batch:
        cur.executemany(INSERT_SQL, batch)
        inserted += len(batch)

    return inserted


def import_format_b(ws, province: str, cur, existing: set) -> int:
    """导入格式B（旧版按年文件）"""
    rows = ws.iter_rows(values_only=True)
    header = [safe_str(h) for h in next(rows)]
    col = {h: i for i, h in enumerate(header)}

    def get(row, name):
        idx = col.get(name)
        return row[idx] if idx is not None else None

    inserted = 0
    batch = []
    for row in rows:
        year = safe_int(get(row, '年份'))
        school = safe_str(get(row, '学校') or get(row, '院校名称'))
        major = safe_str(get(row, '专业'))
        min_rank = safe_int(get(row, '最低分位次') or get(row, '最低分位次(含专科)'))

        if not year or not school or not major or not min_rank:
            continue

        key = (province, year, school, major)
        if key in existing:
            continue
        existing.add(key)

        school_prov = safe_str(get(row, '省份') or get(row, '学校方向') or get(row, '学校所在'))
        # 旧格式省份可能带前缀如 "H湖北" → 清理
        if school_prov and len(school_prov) > 2 and school_prov[0].isalpha() and school_prov[1].isupper():
            school_prov = school_prov[1:]

        batch.append((
            safe_str(get(row, '招生代码') or get(row, '院校代码')),
            school,
            major,
            '',  # major_group
            province,
            year,
            safe_str(get(row, '科目') or get(row, '科类') or get(row, '批次')),
            safe_str(get(row, '选科要求')),
            safe_int(get(row, '最低分') or get(row, '最低分数')),
            min_rank,
            safe_int(get(row, '录取人数')),
            school_prov,
            safe_str(get(row, '学校性质')),
            safe_str(get(row, '是否985')),
            safe_str(get(row, '是否211')),
        ))

        if len(batch) >= 2000:
            cur.executemany(INSERT_SQL, batch)
            inserted += len(batch)
            batch = []

    if batch:
        cur.executemany(INSERT_SQL, batch)
        inserted += len(batch)

    return inserted


INSERT_SQL = """
INSERT OR IGNORE INTO admission_records
  (school_code, school_name, major_name, major_group, province, year, batch,
   subject_req, min_score, min_rank, admit_count, school_province, school_nature, is_985, is_211)
VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
"""

# ── 省份 → 文件路径配置 ───────────────────────────────────────────
# 格式A省份（有统一22-25文件）
PROVINCES_FORMAT_A = {
    "重庆":  "03、重庆-2026志愿填报资料【永久更新】/2、重庆录取数据22-25【持续更新】/22-25年全国高校在重庆的专业录取分数.xlsx",
    "上海":  "05、上海-2026志愿填报资料【永久更新】/2、上海录取数据22-25【持续更新】/22-25年全国高校在上海的专业录取分数.xlsx",
    "内蒙古":"07、内蒙古-2026志愿填报资料【永久更新】/2、内蒙古录取数据22-25【持续更新】/22-25年全国高校在内蒙古的专业录取分数.xlsx",
    "安徽":  "10、安徽-2026志愿填报资料【永久更新】/2、安徽高考录取数据22-25【持续更新】/22-25年全国高校在安徽的专业录取分数.xlsx",
    "江西":  "11、江西-2026志愿填报资料包【永久更新】/2、江西高考录取数据22-25【持续更新】/22-25年全国高校在江西的专业录取分数.xlsx",
    "河北":  "13、河北-2026志愿填报资料【永久更新】/2、河北录取数据22-25【持续更新】/22-25年全国高校在河北的专业录取分数.xlsx",
    "天津":  "15、天津-2026志愿填报资料【永久更新】/2、天津高考录取数据22-25【持续更新】/22-25年全国高校在天津的专业录取分数.xlsx",
    "吉林":  "17、吉林-2026志愿填报资料【永久更新】/2、吉林高考录取数据22-25【持续更新】/22-25年全国高校在吉林的专业录取分数.xlsx",
    "广西":  "19、广西-2026志愿填报资料【永久更新】/2、广西高考数据22-25【持续更新】/22-25年全国高校在广西的专业录取分数.xlsx",
    "云南":  "21、云南-2026志愿填报资料【永久更新】/2、云南高考录取数据22-25【持续更新】/22-25年全国高校在云南的专业录取分数.xlsx",
    "四川":  "23、四川-2026志愿填报资料【永久更新】/2、四川高考录取数据22-25【持续更新】/22-25年全国高校在四川的专业录取分数.xlsx",
    "陕西":  "25、陕西-2026志愿填报资料【永久更新】/2、陕西高考录取数据22-25【持续更新】/22-25年全国高校在陕西的专业录取分数.xlsx",
    "山西":  "26、山西-2026志愿填报资料【永久更新】/2、山西高考录取数据22-25【持续更新】/22-25年全国高校在山西的专业录取分数.xlsx",
    "青海":  "28、青海-2026志愿填报资料【永久更新】/2、青海高考录取数据22-25【持续更新】/22-25年全国高校在青海的专业录取分数.xlsx",
    "新疆":  "29、新疆-2026志愿填报资料【永久更新】/2、新疆高考录取数据22-25年【持续更新】/22-25年全国高校在新疆的专业录取分数.xlsx",
    "西藏":  "31、西藏-2026志愿填报资料【永久更新】/2、西藏高考数据22-25【持续更新】/22-25年全国高校在西藏的专业录取分数.xlsx",
}

# 格式B省份（旧版，按年文件）
PROVINCES_FORMAT_B = {
    "湖北": "16、湖北-2026志愿填报资料【永久更新】/1、湖北高考录取数据17-24/湖北_专业分数线",
    "湖南": "02、湖南-2026志愿填报资料【永久更新】/1、湖南高考录取数据17-24/湖南_专业分数线",
}

def main():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    cur = conn.cursor()

    # 已有数据的去重集合（内存）
    print("加载已有记录键（去重）...")
    cur.execute("SELECT province, year, school_name, major_name FROM admission_records")
    existing = set(cur.fetchall())
    print(f"  已有记录：{len(existing):,} 条")

    total_inserted = 0

    # ── 格式A ──
    for province, rel_path in PROVINCES_FORMAT_A.items():
        file_path = os.path.join(DATA_ROOT, rel_path)
        if not os.path.exists(file_path):
            print(f"⚠️  {province}: 文件不存在 {file_path}")
            continue

        t0 = time.time()
        print(f"📥 导入 {province}（格式A）...", end='', flush=True)
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            n = import_format_a(ws, province, cur, existing)
            wb.close()
            conn.commit()
            total_inserted += n
            print(f" +{n:,} 条  ({time.time()-t0:.1f}s)")
        except Exception as e:
            print(f" ❌ 错误: {e}")
            conn.rollback()

    # ── 格式B ──
    for province, rel_dir in PROVINCES_FORMAT_B.items():
        dir_path = os.path.join(DATA_ROOT, rel_dir)
        if not os.path.exists(dir_path):
            print(f"⚠️  {province}: 目录不存在 {dir_path}")
            continue

        files = sorted(glob.glob(os.path.join(dir_path, "*.xlsx")))
        n_total = 0
        t0 = time.time()
        print(f"📥 导入 {province}（格式B，{len(files)}个文件）...", end='', flush=True)
        for f in files:
            if "专科" in f or "艺" in f:
                continue  # 跳过专科和艺考
            try:
                wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
                ws = wb.active
                n = import_format_b(ws, province, cur, existing)
                wb.close()
                n_total += n
            except Exception as e:
                print(f"\n  {os.path.basename(f)}: {e}", end='')

        conn.commit()
        total_inserted += n_total
        print(f" +{n_total:,} 条  ({time.time()-t0:.1f}s)")

    # ── 统计 ──
    print(f"\n✅ 导入完成！新增 {total_inserted:,} 条记录")
    cur.execute("SELECT province, COUNT(*), MIN(year), MAX(year) FROM admission_records GROUP BY province ORDER BY COUNT(*) DESC")
    print("\n📊 各省数据覆盖：")
    for row in cur.fetchall():
        print(f"  {row[0]:8s}: {row[1]:>8,}条  ({row[2]}-{row[3]})")

    conn.close()

if __name__ == "__main__":
    main()
