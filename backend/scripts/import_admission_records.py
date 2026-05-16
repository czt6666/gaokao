#!/usr/bin/env python3
"""
全国31省2022-2025高考专业录取数据导入脚本

功能:
1. 清空 admission_records 表，从0开始导入
2. 读取 province_22-25_files.json 中的31省Excel
3. 自动识别列名（支持列顺序不同、列名差异）
4. 解析选科要求 -> subject_must, subject_any_of
5. 解析批次 -> batch_type
6. 解析专业备注 -> major_restrictions
7. 批量写入SQLite，每3000行提交一次

使用方式:
  # 1. 先执行迁移（增加字段）
  python backend/migrations/migrate_add_import_fields.py --db backend/gaokao.db

  # 2. 执行导入（正式）
  python backend/scripts/import_admission_records.py --db backend/gaokao.db --mapping province_22-25_files.json

  # 3. 仅预览列映射，不写入
  python backend/scripts/import_admission_records.py --db backend/gaokao.db --mapping province_22-25_files.json --dry-run

原则:
  - 准确第一: 必要列缺失则跳过整个省份
  - 兼容差异: 浙江/宁夏15列、新疆"最低分位"、列顺序不同均可自动适配
  - 容错处理: 单条行解析失败仅跳过该行，不影响整体导入
  - 全量写入: 31省全部读完才结束，最后统计各省行数
"""
import argparse
import json
import os
import sqlite3
import sys
import time
from typing import Optional

import openpyxl

# 把 backend/scripts 加入路径，以便导入同目录模块
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from batch_type_map import get_batch_type
from subject_rule_map import parse_subject_fields
from major_note_expr import NoteExpressionEngine


# ═══════════════════════════════════════════════════════════════
# 列名映射配置
# 支持同一含义的多种列名变体，按关键词包含匹配
# ═══════════════════════════════════════════════════════════════

COLUMN_ALIASES = {
    "year":            ["年份"],
    "school_code":     ["院校代码"],
    "school_name":     ["院校名称"],
    "major_name":      ["专业"],
    "major_group":     ["所属专业组", "专业组"],
    "batch":           ["批次"],
    "subject_req":     ["选科要求"],
    "admit_count":     ["录取人数", "招生人数"],
    "min_score":       ["最低分数", "最低分"],
    "min_rank":        ["最低位次", "最低分位"],
    "school_province": ["学校所在", "所在省"],
    "school_nature":   ["学校性质", "公私性质"],
    "is_985":          ["是否985"],
    "is_211":          ["是否211"],
    "major_note":      ["专业备注"],
}

# 必要列：缺少任意一个则跳过该省份
REQUIRED_FIELDS = ["year", "school_name", "major_name", "batch"]

# 可选列：缺少时仅打印警告，不影响导入
OPTIONAL_FIELDS = [
    "min_score", "min_rank", "admit_count", "school_province",
    "school_nature", "is_985", "is_211", "major_note", "major_group",
    "school_code", "subject_req",
]


# ═══════════════════════════════════════════════════════════════
# 专业备注解析引擎（全局单例，避免重复初始化）
# ═══════════════════════════════════════════════════════════════

_NOTE_ENGINE = NoteExpressionEngine()


# ═══════════════════════════════════════════════════════════════
# 工具函数
# ═══════════════════════════════════════════════════════════════

def find_col_index(headers: list[str], aliases: list[str]) -> Optional[int]:
    """在表头中查找匹配的列索引（关键词包含匹配）"""
    for alias in aliases:
        for idx, h in enumerate(headers):
            if h and alias in h:
                return idx
    return None


def build_col_map(headers: list[str]) -> dict[str, int]:
    """根据表头构建 字段名->列索引 映射"""
    mapping = {}
    for field, aliases in COLUMN_ALIASES.items():
        idx = find_col_index(headers, aliases)
        if idx is not None:
            mapping[field] = idx
    return mapping


def safe_int(val, default: Optional[int] = None) -> Optional[int]:
    """安全转整数。Excel中常见 '620.0' 或空值"""
    if val is None:
        return default
    try:
        # 先转 float 再转 int，兼容 "620.0"
        return int(float(val))
    except (ValueError, TypeError, OverflowError):
        return default


def safe_str(val, default: str = "") -> str:
    """安全转字符串并清理"""
    if val is None:
        return default
    s = str(val).strip()
    # Excel 中常见的无意义占位符
    if s in ("-", "—", "None", "nan", "NULL"):
        return default
    return s


def parse_major_restrictions(note: str) -> str:
    """解析专业备注，返回限制标签串"""
    if not note:
        return ""
    return _NOTE_ENGINE.evaluate(note)


def get_cell(row: tuple, col_map: dict[str, int], field: str, default=None):
    """安全地从行数据中按字段名取值"""
    idx = col_map.get(field)
    if idx is None:
        return default
    return row[idx]


# ═══════════════════════════════════════════════════════════════
# 单省份导入
# ═══════════════════════════════════════════════════════════════

def import_province(conn: sqlite3.Connection, path: str, province: str) -> tuple[int, int]:
    """
    导入单个省份的Excel。
    返回: (成功写入行数, 跳过的错误/空行数)
    """
    if not os.path.exists(path):
        print(f"  [SKIP] {province}: 文件不存在 {path}")
        return 0, 0

    print(f"  [OPEN] {province}: {path}")

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        print(f"  [ERROR] {province}: 无法打开文件: {e}")
        return 0, 0

    # 读取表头
    raw_headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    headers = [str(h).strip() if h else "" for h in raw_headers]
    col_map = build_col_map(headers)

    # 检查必要列
    missing_required = [f for f in REQUIRED_FIELDS if f not in col_map]
    if missing_required:
        print(f"  [ERROR] {province}: 缺少必要列 {missing_required}，可用列: {headers}")
        wb.close()
        return 0, 0

    # 检查可选列缺失情况
    missing_optional = [f for f in OPTIONAL_FIELDS if f not in col_map]
    if missing_optional:
        print(f"  [WARN] {province}: 缺少可选列 {missing_optional}")

    # 预编译插入SQL（19个字段，与当前表结构一致）
    sql = """
    INSERT INTO admission_records
    (school_code, school_name, major_name, major_group, province, year, batch,
     subject_req, min_score, min_rank, admit_count, school_province, school_nature,
     is_985, is_211, batch_type, subject_must, subject_any_of, major_restrictions)
    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """

    cur = conn.cursor()
    inserted = 0
    skipped = 0
    buffer: list[tuple] = []
    BATCH_SIZE = 3000

    for row in ws.iter_rows(min_row=2, values_only=True):
        # 跳过全空行
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue

        try:
            year = safe_int(row[col_map["year"]])
            if not year:
                skipped += 1
                continue

            school_name = safe_str(row[col_map["school_name"]])
            major_name = safe_str(row[col_map["major_name"]])
            if not school_name or not major_name:
                skipped += 1
                continue

            batch = safe_str(row[col_map["batch"]])
            subject_req_raw = safe_str(get_cell(row, col_map, "subject_req"))

            # 调用映射模块
            subject_must, subject_any_of = parse_subject_fields(subject_req_raw)
            batch_type = get_batch_type(batch)

            # 专业备注解析
            major_note = safe_str(get_cell(row, col_map, "major_note"))
            major_restrictions = parse_major_restrictions(major_note)

            record = (
                safe_str(get_cell(row, col_map, "school_code")),
                school_name,
                major_name,
                safe_str(get_cell(row, col_map, "major_group")),
                province,
                year,
                batch,
                subject_req_raw,
                safe_int(get_cell(row, col_map, "min_score")),
                safe_int(get_cell(row, col_map, "min_rank")),
                safe_int(get_cell(row, col_map, "admit_count")),
                safe_str(get_cell(row, col_map, "school_province")),
                safe_str(get_cell(row, col_map, "school_nature")),
                safe_str(get_cell(row, col_map, "is_985")),
                safe_str(get_cell(row, col_map, "is_211")),
                batch_type,
                subject_must,
                subject_any_of,
                major_restrictions,
            )
            buffer.append(record)

            if len(buffer) >= BATCH_SIZE:
                cur.executemany(sql, buffer)
                conn.commit()
                inserted += len(buffer)
                buffer = []

        except Exception as e:
            skipped += 1
            if skipped <= 5:
                # 只打印前5条错误详情，避免刷屏
                print(f"  [ERROR] {province} 行解析失败: {e}")
            continue

    # 写入剩余缓冲
    if buffer:
        cur.executemany(sql, buffer)
        conn.commit()
        inserted += len(buffer)

    wb.close()
    print(f"  [DONE] {province}: 写入 {inserted} 行, 跳过 {skipped} 行")
    return inserted, skipped


# ═══════════════════════════════════════════════════════════════
# 主流程
# ═══════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="Import 31-province admission records")
    parser.add_argument("--db", default="backend/gaokao.db", help="SQLite database path")
    parser.add_argument("--mapping", default="province_22-25_files.json", help="Province file mapping JSON")
    parser.add_argument("--dry-run", action="store_true", help="Preview column mapping only, do not write")
    args = parser.parse_args()

    # 读取省份映射
    with open(args.mapping, "r", encoding="utf-8") as f:
        provinces = json.load(f)

    conn = sqlite3.connect(args.db)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    conn.execute("PRAGMA cache_size=-200000")  # 200MB page cache

    if args.dry_run:
        print("[DRY RUN] 仅预览列映射，不写入数据")
    else:
        print("[CLEAR] 清空 admission_records 表...")
        conn.execute("DELETE FROM admission_records")
        # 重置自增ID（如果存在 sqlite_sequence）
        try:
            conn.execute("DELETE FROM sqlite_sequence WHERE name='admission_records'")
        except Exception:
            pass
        conn.commit()
        print("[CLEAR] 完成")

    total_inserted = 0
    total_skipped = 0
    start_time = time.time()

    for item in provinces:
        prov = item["province_name"]
        path = item["full_path"]
        print(f"\n[IMPORT] {prov}")
        if args.dry_run:
            # 干跑模式下仅打印列映射
            if not os.path.exists(path):
                print(f"  [SKIP] 文件不存在")
                continue
            try:
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                ws = wb.active
                raw_headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                headers = [str(h).strip() if h else "" for h in raw_headers]
                col_map = build_col_map(headers)
                print(f"  [DRY] 列数={len(headers)}, 映射={col_map}")
                missing = [f for f in REQUIRED_FIELDS if f not in col_map]
                if missing:
                    print(f"  [DRY] 缺少必要列: {missing}")
                else:
                    print(f"  [DRY] 必要列全部匹配")
                wb.close()
            except Exception as e:
                print(f"  [ERROR] 无法预览: {e}")
            continue

        n, s = import_province(conn, path, prov)
        total_inserted += n
        total_skipped += s

    conn.close()

    elapsed = time.time() - start_time
    print(f"\n{'='*60}")
    print(f"总计: 写入 {total_inserted} 行, 跳过 {total_skipped} 行")
    print(f"耗时: {elapsed:.1f}s ({elapsed/60:.1f}min)")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
