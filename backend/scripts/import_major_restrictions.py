#!/usr/bin/env python3
"""
从31省Excel的"专业备注"列提取限制标签，更新到 admission_records 表。

使用方式：
  python import_major_restrictions.py --db backend/gaokao.db --data-dir "E:\高考程序素材\05、【琢玉攻略】高考数据"
"""
import argparse
import json
import os
import sqlite3

import openpyxl


RULES = {
    "gender": {
        "male_only": ["只招男生", "限男生", "招男生", "（男）（", "(男)(", "（男）", "(男)", "男生"],
        "female_only": ["只招女生", "限女生", "招女生", "（女）（", "(女)(", "（女）", "(女)", "女生"],
    },
    "health": {
        "color_blind": [
            "不招色盲", "色盲限报", "色盲不予录取", "色盲不录", "色盲受限",
            "色盲、色弱不录", "色盲、色弱限报", "色盲、色弱考生限报",
            "色盲色弱限报", "色盲色弱受限", "无色盲", "不招单色识别不全"
        ],
        "color_weak": ["不招色弱", "色弱限报", "色弱不录", "色弱受限", "无色弱"],
        "monochrome": ["不招单色识别不全", "单色识别不全", "单色识别能力异常"],
    },
    "lang": {
        "english_only": [
            "只招英语", "招英语", "外语语种英语", "语种:英语", "语种：英语",
            "语种为英语", "外语语种:英语", "外语语种：英语", "外语语种要求英语",
            "语种要求英语", "外语要求:英语", "外语要求：英语"
        ],
        "english_japanese": [
            "只招英语、日语", "招英语、日语", "外语语种:英日",
            "语种:英语,日语", "语种：英语，日语", "只招英语，日语"
        ],
        "english_russian": [
            "只招英语、俄语", "招英语、俄语", "语种:英语,俄语", "语种：英语，俄语"
        ],
    },
    "special": {
        "experiment": [
            "实验班", "基地班", "拔尖班", "卓越班", "创新班", "菁英班", "英才班",
            "大师班", "阶平班", "韬奋班", "自强班", "梁希班", "华佗班", "岐黄班",
            "屠呦呦班", "启明班", "成思危班", "侯宗濂班", "钱学森班", "吕振羽班",
            "正谊明道班"
        ],
        "national_special": ["国家专项", "国家专项计划"],
        "local_special": ["地方专项", "地方专项计划"],
        "oriented": ["定向", "定向培养", "定向西藏"],
        "free_teacher": ["公费师范", "师范类", "（师范）", "(师范)", "师范"],
        "sino_foreign": [
            "中外合作", "中美合作", "中英合作", "中法合作", "中德合作", "中加合作", "中马合作"
        ],
    },
}


def extract_restrictions(note: str) -> dict:
    """从专业备注文本中提取限制标签。"""
    if not note or not isinstance(note, str):
        return {}
    note = note.strip()
    if not note:
        return {}

    result = {}
    for category, rules in RULES.items():
        matched = False
        for label, keywords in rules.items():
            for kw in keywords:
                if kw in note:
                    result[category] = label
                    matched = True
                    break
            if matched:
                break
    return result


def format_restrictions(res: dict) -> str:
    """将限制字典格式化为逗号分隔字符串。"""
    if not res:
        return ""
    tags = [f"{k}:{v}" for k, v in res.items()]
    return ",".join(tags)


def get_headers(ws) -> dict:
    """读取表头并返回列名->索引映射。"""
    headers = {}
    for i, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1))):
        if cell.value:
            headers[str(cell.value).strip()] = i
    return headers


def import_province(conn: sqlite3.Connection, path: str, province: str) -> int:
    """导入单个省份的Excel，更新 admission_records 的 major_restrictions。"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = get_headers(ws)

    # 定位关键列
    year_col = headers.get("年份")
    school_col = headers.get("院校名称")
    major_col = headers.get("专业")
    note_col = headers.get("专业备注")

    if year_col is None or school_col is None or major_col is None:
        print(f"  [WARN] {province}: 缺少关键列，跳过")
        wb.close()
        return 0

    cur = conn.cursor()
    updated = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) <= max(filter(None, [year_col, school_col, major_col, note_col])):
            continue

        year = row[year_col]
        school = str(row[school_col]).strip() if row[school_col] else ""
        major = str(row[major_col]).strip() if row[major_col] else ""

        if not year or not school or not major:
            continue

        # 提取限制标签
        note = str(row[note_col]).strip() if note_col is not None and row[note_col] else ""
        restrictions = format_restrictions(extract_restrictions(note))

        # 更新数据库
        cur.execute(
            "UPDATE admission_records SET major_restrictions = ? WHERE province = ? AND year = ? AND school_name = ? AND major_name = ?",
            (restrictions, province, int(year), school, major),
        )
        if cur.rowcount > 0:
            updated += cur.rowcount

        if updated % 2000 == 0:
            conn.commit()

    conn.commit()
    wb.close()
    return updated


def main():
    parser = argparse.ArgumentParser(description="Import major restrictions from Excel to database")
    parser.add_argument("--db", default="backend/gaokao.db", help="SQLite database path")
    parser.add_argument("--data-dir", default="E:\\高考程序素材\\05、【琢玉攻略】高考数据", help="Root directory of Excel files")
    args = parser.parse_args()

    # 读取省份文件映射
    with open("province_22-25_files.json", "r", encoding="utf-8") as f:
        files = json.load(f)

    conn = sqlite3.connect(args.db)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")

    total_updated = 0
    for item in files:
        prov = item["province_name"]
        path = item["full_path"]
        if not os.path.exists(path):
            print(f"[SKIP] {prov}: file not found {path}")
            continue
        print(f"[IMPORT] {prov} ...", end="", flush=True)
        n = import_province(conn, path, prov)
        total_updated += n
        print(f" updated {n} rows")

    conn.close()
    print(f"\n[Done] Total updated: {total_updated} rows")


if __name__ == "__main__":
    main()
