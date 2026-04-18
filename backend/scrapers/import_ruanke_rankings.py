"""
Import 2025软科中国大学专业排名 into subject_evaluations table.
This adds school-major quality ratings (A+/A/B+/B/C+/C) for 30,931 combinations.
These enhance the recommendation engine's quality scoring for majors.

Usage:
  python import_ruanke_rankings.py
"""
import os
import sys
import sqlite3
import openpyxl

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB_PATH = os.path.join(BASE, "gaokao.db")
DATA_DIR = os.path.expanduser("~/Desktop/高考程序素材")

RANKINGS_FILE = os.path.join(
    DATA_DIR,
    "00、志愿填报必备资料",
    "1、大学排名",
    "2025软科中国大学专业排名Excel版.xlsx",
)

# Also import 2024软科中国大学领域评级结果.xlsx if it has useful data
DOMAIN_RATINGS_FILE = os.path.join(
    DATA_DIR,
    "00、志愿填报必备资料",
    "1、大学排名",
    "22-25软科排名",
    "2024软科中国大学领域评级结果.xlsx",
)


def ensure_table(conn: sqlite3.Connection):
    conn.execute("""
        CREATE TABLE IF NOT EXISTS subject_evaluations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            school_name VARCHAR,
            school_code VARCHAR,
            subject_code VARCHAR,
            subject_name VARCHAR,
            grade VARCHAR,
            category VARCHAR,
            major_category VARCHAR
        )
    """)
    conn.commit()


def import_ruanke_major_rankings(conn: sqlite3.Connection) -> int:
    """Import 2025软科中国大学专业排名 into subject_evaluations."""
    if not os.path.exists(RANKINGS_FILE):
        print(f"[WARN] File not found: {RANKINGS_FILE}")
        return 0

    print(f"Reading {os.path.basename(RANKINGS_FILE)} ...")
    wb = openpyxl.load_workbook(RANKINGS_FILE, read_only=True, data_only=True)
    ws = wb.active

    headers = [str(c) if c else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    col = {h: i for i, h in enumerate(headers) if h}

    inserted = 0
    skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue

        def get(name):
            idx = col.get(name)
            return row[idx] if idx is not None else None

        school_name = str(get("学校名称") or "").strip()
        major_name  = str(get("专业名称") or "").strip()
        major_cat   = str(get("专业类") or "").strip()
        category    = str(get("门类") or "").strip()
        grade       = str(get("评级") or "").strip()
        major_code  = str(get("专业代码") or "").strip()

        if not school_name or not major_name or not grade:
            skipped += 1
            continue

        try:
            conn.execute("""
                INSERT OR IGNORE INTO subject_evaluations
                    (school_name, school_code, subject_code, subject_name, grade, category, major_category)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (school_name, None, major_code, major_name, grade, "软科专业排名2025", major_cat))
            inserted += 1
        except Exception as e:
            skipped += 1

        if inserted % 5000 == 0:
            conn.commit()
            print(f"  ... {inserted} rows inserted")

    conn.commit()
    wb.close()
    return inserted


def import_ruanke_school_rankings(conn: sqlite3.Connection) -> int:
    """Import 2025年软科中国大学排名 to update schools.rank_2025."""
    rankings_file = os.path.join(
        DATA_DIR,
        "00、志愿填报必备资料",
        "1、大学排名",
        "22-25软科排名",
        "2025年软科中国大学排名数据.xlsx",
    )
    if not os.path.exists(rankings_file):
        print(f"[WARN] File not found: {rankings_file}")
        return 0

    print(f"Reading {os.path.basename(rankings_file)} ...")
    wb = openpyxl.load_workbook(rankings_file, read_only=True)
    ws = wb.active

    headers = [str(c) if c else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    col = {h: i for i, h in enumerate(headers) if h}

    updated = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue

        def get(name):
            idx = col.get(name)
            return row[idx] if idx is not None else None

        rank = get("排名")
        name = str(get("高校名称") or "").strip()

        if not rank or not name:
            continue

        try:
            rank_int = int(rank)
            cursor = conn.execute(
                "UPDATE schools SET rank_2025 = ? WHERE name = ? AND (rank_2025 IS NULL OR rank_2025 = 0)",
                (rank_int, name)
            )
            if cursor.rowcount > 0:
                updated += 1
        except (ValueError, TypeError):
            pass

    conn.commit()
    wb.close()
    return updated


def main():
    conn = sqlite3.connect(DB_PATH)
    ensure_table(conn)

    # Check current count
    cur = conn.execute("SELECT COUNT(*) FROM subject_evaluations WHERE category = '软科专业排名2025'")
    existing = cur.fetchone()[0]
    if existing > 0:
        print(f"Already imported {existing} 软科专业排名 records. Skipping (use DELETE to reimport).")
    else:
        n = import_ruanke_major_rankings(conn)
        print(f"✓ Imported {n:,} school-major rankings into subject_evaluations")

    # Update school rankings
    n2 = import_ruanke_school_rankings(conn)
    print(f"✓ Updated {n2} school rank_2025 values")

    # Final counts
    cur = conn.execute("SELECT category, COUNT(*) FROM subject_evaluations GROUP BY category")
    print("\n=== subject_evaluations by category ===")
    for cat, cnt in cur.fetchall():
        print(f"  {cat or '(null)'}: {cnt:,}")

    conn.close()


if __name__ == "__main__":
    main()
