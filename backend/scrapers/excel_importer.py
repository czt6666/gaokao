#!/usr/bin/env python3
"""
excel_importer.py — Import locally purchased gaokao historical Excel data into SQLite.

Supported sources:
  beijing_major   — 北京专业分数线 2017–2021 → admission_records
  beijing_control — 北京投档线 2017–2021    → admission_records
  national_control — 全国批次线 2014–2023   → province_control_lines

Usage examples:
  python excel_importer.py --source all
  python excel_importer.py --source beijing_major
  python excel_importer.py --source national_control
  python excel_importer.py --status
  python excel_importer.py --source all --data-dir /custom/path/to/素材
"""

import argparse
import os
import sqlite3
import sys

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Path defaults
# ---------------------------------------------------------------------------

DEFAULT_DATA_DIR = "/Users/Admin/Desktop/高考程序素材"
DEFAULT_DB_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "..", "gaokao.db"
)

# ---------------------------------------------------------------------------
# Batch-type classification helpers
# ---------------------------------------------------------------------------

SKIP_KEYWORDS   = ["艺术", "体育", "飞行", "国防生"]
EARLY_KEYWORDS  = ["提前批", "强基", "专项"]
JUNIOR_KEYWORDS = ["专科", "高职"]


def classify_batch_type(batch: str, major_name: str, edu_level: str) -> str:
    """Return 'skip' | 'early' | 'junior' | 'normal' for a row."""
    batch_str     = str(batch or "")
    major_str     = str(major_name or "")
    edu_str       = str(edu_level or "")

    # Skip: art, sport, flight, defence
    for kw in SKIP_KEYWORDS:
        if kw in batch_str:
            return "skip"

    # Junior: edu level explicitly says 专科, or batch/major contains junior keywords
    if "专科" in edu_str:
        return "junior"
    for kw in JUNIOR_KEYWORDS:
        if kw in batch_str or kw in major_str:
            return "junior"

    # Early: 提前批, 强基, 专项
    for kw in EARLY_KEYWORDS:
        if kw in batch_str:
            return "early"

    return "normal"


# ---------------------------------------------------------------------------
# Value conversion helpers
# ---------------------------------------------------------------------------

def to_int(value) -> int | None:
    """Convert cell value to int, returning None for blanks/dashes/zeros."""
    if value is None:
        return None
    s = str(value).strip()
    if s in ("", "-", "—", "－"):
        return None
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None


def to_str(value) -> str | None:
    """Return stripped string or None."""
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def is_yes(value) -> str:
    """Normalise 是/否 flags to '是'/'否'."""
    s = str(value or "").strip()
    return "是" if s == "是" else "否"


# ---------------------------------------------------------------------------
# Database helpers
# ---------------------------------------------------------------------------

def get_connection(db_path: str) -> sqlite3.Connection:
    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    return conn


def ensure_batch_type_column(conn: sqlite3.Connection) -> None:
    """Add batch_type column to admission_records if it doesn't exist."""
    cursor = conn.execute("PRAGMA table_info(admission_records)")
    cols = [row[1] for row in cursor.fetchall()]
    if "batch_type" not in cols:
        conn.execute("ALTER TABLE admission_records ADD COLUMN batch_type VARCHAR")
        conn.commit()
        print("  [schema] Added batch_type column to admission_records.")


def ensure_province_control_lines_table(conn: sqlite3.Connection) -> None:
    """Create province_control_lines if it doesn't exist."""
    conn.execute("""
        CREATE TABLE IF NOT EXISTS province_control_lines (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            province    VARCHAR,
            year        INTEGER,
            batch       VARCHAR,
            subject_type VARCHAR,
            score       INTEGER
        )
    """)
    conn.commit()


# ---------------------------------------------------------------------------
# Source 1: 北京专业分数线
# ---------------------------------------------------------------------------

BEIJING_MAJOR_YEARS = [2017, 2018, 2019, 2020, 2021]

INSERT_ADMISSION = """
    INSERT OR IGNORE INTO admission_records
        (school_code, school_name, major_name, major_group,
         province, year, batch, subject_req,
         min_score, min_rank, admit_count,
         school_province, school_nature, is_985, is_211, batch_type)
    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
"""


def import_beijing_major(conn: sqlite3.Connection, data_dir: str) -> dict:
    """Import 北京专业分数线 2017–2021."""
    folder = os.path.join(
        data_dir,
        "18、北京-2026志愿填报资料【永久更新】",
        "1、北京高考录取数据17-24",
        "北京_专业分数线",
    )
    ensure_batch_type_column(conn)

    summary = {}
    for year in BEIJING_MAJOR_YEARS:
        filename = f"北京_专业分数线_{year}.xlsx"
        filepath = os.path.join(folder, filename)
        if not os.path.exists(filepath):
            print(f"  [WARN] File not found, skipping: {filepath}")
            continue

        print(f"  Reading {filename} …")
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active

        # Read headers from row 1
        headers = [to_str(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        col = {h: i for i, h in enumerate(headers) if h}

        inserted = 0
        skipped  = 0
        row_num  = 1

        for row in ws.iter_rows(min_row=2, values_only=True):
            row_num += 1
            if all(v is None for v in row):
                continue

            def get(name):
                idx = col.get(name)
                return row[idx] if idx is not None else None

            year_val    = to_int(get("年份")) or year
            school      = to_str(get("学校"))
            major       = to_str(get("专业")) or "[院校最低分]"
            min_score   = to_int(get("最低分"))
            min_rank    = to_int(get("最低分排名"))
            subject_req = to_str(get("科类"))
            batch       = to_str(get("批次"))
            school_code = to_int(get("全国统一招生代码"))
            school_prov = to_str(get("省份"))
            school_nat  = to_str(get("办学性质"))
            is_985      = is_yes(get("_985"))
            is_211      = is_yes(get("_211"))
            edu_level   = to_str(get("学历类别"))

            # Skip rows with no usable score or rank
            if (min_score is None or min_score == 0) and (min_rank is None or min_rank == 0):
                skipped += 1
                continue

            # Classify batch
            batch_type = classify_batch_type(batch, major, edu_level)
            if batch_type == "skip":
                skipped += 1
                continue

            conn.execute(INSERT_ADMISSION, (
                str(school_code) if school_code else None,
                school,
                major,
                None,           # major_group not in this file
                "北京",         # province (生源地)
                year_val,
                batch,
                subject_req,
                min_score,
                min_rank if min_rank and min_rank != 0 else None,
                None,           # admit_count not in this file
                school_prov,
                school_nat,
                is_985,
                is_211,
                batch_type,
            ))
            inserted += 1

            if inserted % 1000 == 0:
                conn.commit()
                print(f"    … {inserted} rows inserted so far (row {row_num})")

        conn.commit()
        wb.close()
        summary[filename] = inserted
        print(f"  {filename}: inserted {inserted}, skipped {skipped}")

    return summary


# ---------------------------------------------------------------------------
# Source 2: 北京投档线
# ---------------------------------------------------------------------------

BEIJING_CONTROL_YEARS = [2017, 2018, 2019, 2020, 2021, 2022]


def import_beijing_control(conn: sqlite3.Connection, data_dir: str) -> dict:
    """Import 北京投档线 2017–2022 + 2023 into admission_records with major_name='[院校最低分]'."""
    folder = os.path.join(
        data_dir,
        "18、北京-2026志愿填报资料【永久更新】",
        "1、北京高考录取数据17-24",
        "北京_投档线",
    )
    ensure_batch_type_column(conn)

    summary = {}
    for year in BEIJING_CONTROL_YEARS:
        filename = f"北京_投档线_{year}.xlsx"
        filepath = os.path.join(folder, filename)
        if not os.path.exists(filepath):
            print(f"  [WARN] File not found, skipping: {filepath}")
            continue

        print(f"  Reading {filename} …")
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active

        headers = [to_str(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        col = {h: i for i, h in enumerate(headers) if h}

        inserted = 0
        skipped  = 0
        row_num  = 1

        for row in ws.iter_rows(min_row=2, values_only=True):
            row_num += 1
            if all(v is None for v in row):
                continue

            def get(name):
                idx = col.get(name)
                return row[idx] if idx is not None else None

            year_val    = to_int(get("年份")) or year
            school      = to_str(get("学校"))
            min_score   = to_int(get("最低分"))
            min_rank    = to_int(get("最低分排名"))
            subject_req = to_str(get("科类"))
            batch       = to_str(get("批次"))
            school_code = to_int(get("全国统一招生代码"))
            school_prov = to_str(get("省份"))
            school_nat  = to_str(get("办学性质"))
            is_985      = is_yes(get("_985"))
            is_211      = is_yes(get("_211"))
            edu_level   = to_str(get("学历类别"))

            # 投档线rows use a fixed major_name
            major = "[院校最低分]"

            # Skip rows with no usable score or rank
            if (min_score is None or min_score == 0) and (min_rank is None or min_rank == 0):
                skipped += 1
                continue

            batch_type = classify_batch_type(batch, major, edu_level)
            if batch_type == "skip":
                skipped += 1
                continue

            conn.execute(INSERT_ADMISSION, (
                str(school_code) if school_code else None,
                school,
                major,
                None,
                "北京",
                year_val,
                batch,
                subject_req,
                min_score,
                min_rank if min_rank and min_rank != 0 else None,
                None,
                school_prov,
                school_nat,
                is_985,
                is_211,
                batch_type,
            ))
            inserted += 1

            if inserted % 1000 == 0:
                conn.commit()
                print(f"    … {inserted} rows inserted so far (row {row_num})")

        conn.commit()
        wb.close()
        summary[filename] = inserted
        print(f"  {filename}: inserted {inserted}, skipped {skipped}")

    # Also import 2023北京投档线.xlsx (different filename, same schema)
    file_2023 = os.path.join(folder, "2023北京投档线.xlsx")
    if os.path.exists(file_2023):
        print(f"  Reading 2023北京投档线.xlsx …")
        try:
            wb23 = openpyxl.load_workbook(file_2023, read_only=True, data_only=True)
            ws23 = wb23["2023高校投档线"] if "2023高校投档线" in wb23.sheetnames else wb23.active
            headers23 = [to_str(cell.value) for cell in next(ws23.iter_rows(min_row=1, max_row=1))]
            col23 = {h: i for i, h in enumerate(headers23) if h}
            ins23, skp23 = 0, 0
            for row in ws23.iter_rows(min_row=2, values_only=True):
                if all(v is None for v in row):
                    continue
                def get23(name):
                    idx = col23.get(name)
                    return row[idx] if idx is not None else None
                year_val    = to_int(get23("年份")) or 2023
                school      = to_str(get23("学校"))
                min_score   = to_int(get23("最低分"))
                min_rank    = to_int(get23("最低分排名"))
                subject_req = to_str(get23("科类"))
                batch       = to_str(get23("批次"))
                school_prov = to_str(get23("省份"))
                school_nat  = to_str(get23("办学性质"))
                if (min_score is None or min_score == 0) and (min_rank is None or min_rank == 0):
                    skp23 += 1
                    continue
                batch_type = classify_batch_type(batch, "[院校最低分]", "")
                if batch_type == "skip":
                    skp23 += 1
                    continue
                conn.execute(INSERT_ADMISSION, (
                    None, school, "[院校最低分]", None, "北京", year_val,
                    batch, subject_req, min_score,
                    min_rank if min_rank and min_rank != 0 else None,
                    None, school_prov, school_nat, None, None, batch_type,
                ))
                ins23 += 1
            conn.commit()
            wb23.close()
            summary["2023北京投档线.xlsx"] = ins23
            print(f"  2023北京投档线.xlsx: inserted {ins23}, skipped {skp23}")
        except Exception as e:
            print(f"  [WARN] Failed to import 2023北京投档线.xlsx: {e}")

    return summary


# ---------------------------------------------------------------------------
# Source 3: 全国批次线 2014–2023
# ---------------------------------------------------------------------------

NATIONAL_CONTROL_FILE = os.path.join(
    "00、志愿填报必备资料",
    "7、全国各省市批次线",
    "2014-2023年各地高考历年分数线(批次线).xlsx",
)

INSERT_CONTROL = """
    INSERT OR IGNORE INTO province_control_lines
        (province, year, batch, subject_type, score)
    VALUES (?, ?, ?, ?, ?)
"""


def import_national_control(conn: sqlite3.Connection, data_dir: str) -> dict:
    """Import 全国批次线 2014–2023 into province_control_lines.

    Only imports rows where 批次 contains '本科' (普通本科批).
    Skips art/sport/专科批 lines.
    """
    ensure_province_control_lines_table(conn)

    filepath = os.path.join(data_dir, NATIONAL_CONTROL_FILE)
    if not os.path.exists(filepath):
        print(f"  [WARN] File not found: {filepath}")
        return {}

    filename = os.path.basename(filepath)
    print(f"  Reading {filename} …")

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    headers = [to_str(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers) if h}

    inserted = 0
    skipped  = 0
    row_num  = 1

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_num += 1
        if all(v is None for v in row):
            continue

        def get(name):
            idx = col.get(name)
            return row[idx] if idx is not None else None

        province     = to_str(get("地区"))
        year         = to_int(get("年份"))
        subject_type = to_str(get("考生类别"))
        batch        = to_str(get("批次"))
        score        = to_int(get("分数线"))

        # Only keep 本科 lines
        if not batch or "本科" not in batch:
            skipped += 1
            continue

        # Skip art/sport/专科 lines even if "本科" appears in the string
        skip = False
        for kw in ["艺术", "体育", "专科", "高职"]:
            if kw in (batch or ""):
                skip = True
                break
        if skip:
            skipped += 1
            continue

        if province is None or year is None or score is None:
            skipped += 1
            continue

        conn.execute(INSERT_CONTROL, (province, year, batch, subject_type, score))
        inserted += 1

        if inserted % 1000 == 0:
            conn.commit()
            print(f"    … {inserted} rows inserted so far (row {row_num})")

    conn.commit()
    wb.close()
    print(f"  {filename}: inserted {inserted}, skipped {skipped}")
    return {filename: inserted}


# ---------------------------------------------------------------------------
# Source 4: 22-25年全国高校在北京的专业录取分数（per-major, 2022-2025）
# ---------------------------------------------------------------------------

def import_beijing_major_2225(conn: sqlite3.Connection, data_dir: str) -> dict:
    """Import 22-25年全国高校在北京的专业录取分数.xlsx into admission_records."""
    filepath = os.path.join(
        data_dir,
        "18、北京-2026志愿填报资料【永久更新】",
        "2、北京高考录取数据22-25【持续更新】",
        "22-25年全国高校在北京的专业录取分数.xlsx",
    )
    if not os.path.exists(filepath):
        print(f"  [WARN] File not found: {filepath}")
        return {}

    ensure_batch_type_column(conn)
    filename = os.path.basename(filepath)
    print(f"  Reading {filename} …")

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    headers = [to_str(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers) if h}

    inserted = skipped = row_num = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_num += 1
        if all(v is None for v in row):
            continue

        def get(name):
            idx = col.get(name)
            return row[idx] if idx is not None else None

        year_val    = to_int(get("年份"))
        school      = to_str(get("院校名称"))
        school_code = to_str(get("院校代码"))
        major       = to_str(get("专业")) or "[院校最低分]"
        batch       = to_str(get("批次")) or ""
        # 科类 + 选科要求 合并为 subject_req
        ke_lei      = to_str(get("科类")) or ""
        xuan_ke     = to_str(get("选科要求")) or ""
        subject_req = xuan_ke if xuan_ke and xuan_ke not in ("不限", "-") else ke_lei
        min_score   = to_int(get("最低分数"))
        min_rank    = to_int(get("最低位次"))
        admit_count = to_int(get("录取人数"))
        school_prov = to_str(get("学校所在"))
        school_nat  = to_str(get("学校性质"))

        if (not min_score or min_score == 0) and (not min_rank or min_rank == 0):
            skipped += 1
            continue

        batch_type = classify_batch_type(batch, major, "")
        if batch_type == "skip":
            skipped += 1
            continue

        conn.execute(INSERT_ADMISSION, (
            school_code, school, major, None,
            "北京", year_val, batch, subject_req,
            min_score, min_rank if min_rank else None,
            admit_count, school_prov, school_nat,
            None, None, batch_type,
        ))
        inserted += 1
        if inserted % 1000 == 0:
            conn.commit()
            print(f"    … {inserted} rows inserted so far (row {row_num})")

    conn.commit()
    wb.close()
    print(f"  {filename}: inserted {inserted}, skipped {skipped}")
    return {filename: inserted}


# ---------------------------------------------------------------------------
# Source 5: 22-25年全国高校在北京的院校录取分数（per-school, 2022-2025）
# ---------------------------------------------------------------------------

def import_beijing_school_2225(conn: sqlite3.Connection, data_dir: str) -> dict:
    """Import 22-25年全国高校在北京的院校录取分数.xlsx into admission_records."""
    filepath = os.path.join(
        data_dir,
        "18、北京-2026志愿填报资料【永久更新】",
        "2、北京高考录取数据22-25【持续更新】",
        "22-25年全国高校在北京的院校录取分数.xlsx",
    )
    if not os.path.exists(filepath):
        print(f"  [WARN] File not found: {filepath}")
        return {}

    ensure_batch_type_column(conn)
    filename = os.path.basename(filepath)
    print(f"  Reading {filename} …")

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    headers = [to_str(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers) if h}

    inserted = skipped = row_num = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_num += 1
        if all(v is None for v in row):
            continue

        def get(name):
            idx = col.get(name)
            return row[idx] if idx is not None else None

        year_val    = to_int(get("年份"))
        school      = to_str(get("院校名称"))
        school_code = to_str(get("院校代码"))
        batch       = to_str(get("批次")) or ""
        ke_lei      = to_str(get("科类")) or ""
        xuan_ke     = to_str(get("选科要求")) or ""
        subject_req = xuan_ke if xuan_ke and xuan_ke not in ("不限", "-") else ke_lei
        min_score   = to_int(get("最低分数"))
        # 该文件列名是"最低分位"
        min_rank    = to_int(get("最低分位"))
        admit_count = to_int(get("录取人数"))
        school_prov = to_str(get("学校所在"))
        school_nat  = to_str(get("学校性质"))
        is_985      = is_yes(get("是否985"))
        is_211      = is_yes(get("是否211"))

        if (not min_score or min_score == 0) and (not min_rank or min_rank == 0):
            skipped += 1
            continue

        batch_type = classify_batch_type(batch, "[院校最低分]", "")
        if batch_type == "skip":
            skipped += 1
            continue

        conn.execute(INSERT_ADMISSION, (
            school_code, school, "[院校最低分]", None,
            "北京", year_val, batch, subject_req,
            min_score, min_rank if min_rank else None,
            admit_count, school_prov, school_nat,
            is_985, is_211, batch_type,
        ))
        inserted += 1
        if inserted % 1000 == 0:
            conn.commit()
            print(f"    … {inserted} rows inserted so far (row {row_num})")

    conn.commit()
    wb.close()
    print(f"  {filename}: inserted {inserted}, skipped {skipped}")
    return {filename: inserted}


# ---------------------------------------------------------------------------
# Source 6: 22-24年省控线汇总表（全国各省2022-2024控制线）
# ---------------------------------------------------------------------------

def import_control_2224(conn: sqlite3.Connection, data_dir: str) -> dict:
    """Import 22-24年省控线汇总表.xlsx into province_control_lines."""
    filepath = os.path.join(
        data_dir,
        "00、志愿填报必备资料",
        "7、全国各省市批次线",
        "22-24年省控线汇总表.xlsx",
    )
    if not os.path.exists(filepath):
        print(f"  [WARN] File not found: {filepath}")
        return {}

    ensure_province_control_lines_table(conn)
    filename = os.path.basename(filepath)
    print(f"  Reading {filename} …")

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    headers = [to_str(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers) if h}

    inserted = skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue

        def get(name):
            idx = col.get(name)
            return row[idx] if idx is not None else None

        province     = to_str(get("省份"))
        year         = to_int(get("年份"))
        batch        = to_str(get("批次/段")) or to_str(get("批次")) or ""
        subject_type = to_str(get("科目")) or to_str(get("考生类别")) or ""
        score        = to_int(get("分数线"))

        # 只保留本科批，跳过艺体专科
        if not batch or "本科" not in batch:
            skipped += 1
            continue
        for kw in ["艺术", "体育", "专科", "高职"]:
            if kw in batch:
                skipped += 1
                break
        else:
            if province and year and score:
                conn.execute(INSERT_CONTROL, (province, year, batch, subject_type, score))
                inserted += 1

    conn.commit()
    wb.close()
    print(f"  {filename}: inserted {inserted}, skipped {skipped}")
    return {filename: inserted}


# ---------------------------------------------------------------------------
# --status report
# ---------------------------------------------------------------------------

def show_status(conn: sqlite3.Connection) -> None:
    """Print record counts in admission_records grouped by province and year."""
    print("\n=== admission_records: counts by province and year ===")
    cursor = conn.execute("""
        SELECT province, year, COUNT(*) as cnt
        FROM admission_records
        GROUP BY province, year
        ORDER BY province, year
    """)
    rows = cursor.fetchall()
    if not rows:
        print("  (no records)")
    else:
        for prov, yr, cnt in rows:
            print(f"  {prov or '?':10s}  {yr or '?'}  {cnt:>7,} records")

    print("\n=== province_control_lines: counts by province and year ===")
    cursor = conn.execute("""
        SELECT province, year, COUNT(*) as cnt
        FROM province_control_lines
        GROUP BY province, year
        ORDER BY province, year
    """)
    rows = cursor.fetchall()
    if not rows:
        print("  (no records)")
    else:
        for prov, yr, cnt in rows:
            print(f"  {prov or '?':10s}  {yr or '?'}  {cnt:>7,} records")


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def parse_args():
    parser = argparse.ArgumentParser(
        description="Import gaokao historical Excel data into SQLite.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "--source",
        choices=["beijing_major", "beijing_control", "national_control",
                 "beijing_major_2225", "beijing_school_2225", "control_2224", "all"],
        default=None,
        help="Which data source to import (default: all when --source not given).",
    )
    parser.add_argument(
        "--status",
        action="store_true",
        help="Show record counts by province/year and exit.",
    )
    parser.add_argument(
        "--db",
        default=DEFAULT_DB_PATH,
        help=f"Path to SQLite database (default: {DEFAULT_DB_PATH}).",
    )
    parser.add_argument(
        "--data-dir",
        default=DEFAULT_DATA_DIR,
        dest="data_dir",
        help=f"Root directory of purchased gaokao data files (default: {DEFAULT_DATA_DIR}).",
    )
    return parser.parse_args()


def main():
    args = parse_args()

    db_path = os.path.abspath(args.db)
    data_dir = os.path.abspath(args.data_dir)

    print(f"Database : {db_path}")
    print(f"Data dir : {data_dir}")

    if not os.path.exists(db_path):
        print(f"ERROR: Database not found at {db_path}")
        sys.exit(1)

    if not os.path.isdir(data_dir):
        print(f"ERROR: Data directory not found at {data_dir}")
        sys.exit(1)

    conn = get_connection(db_path)

    if args.status:
        show_status(conn)
        conn.close()
        return

    # Default to "all" when --source is omitted
    source = args.source or "all"

    all_summary = {}

    if source in ("beijing_major", "all"):
        print("\n[1/3] Importing 北京专业分数线 …")
        summary = import_beijing_major(conn, data_dir)
        all_summary.update(summary)

    if source in ("beijing_control", "all"):
        print("\n[2/3] Importing 北京投档线 …")
        summary = import_beijing_control(conn, data_dir)
        all_summary.update(summary)

    if source in ("national_control", "all"):
        print("\n[3/3] Importing 全国批次线2014-2023 …")
        summary = import_national_control(conn, data_dir)
        all_summary.update(summary)

    if source in ("beijing_major_2225", "all"):
        print("\n[4/6] Importing 北京专业录取分数2022-2025 …")
        summary = import_beijing_major_2225(conn, data_dir)
        all_summary.update(summary)

    if source in ("beijing_school_2225", "all"):
        print("\n[5/6] Importing 北京院校录取分数2022-2025 …")
        summary = import_beijing_school_2225(conn, data_dir)
        all_summary.update(summary)

    if source in ("control_2224", "all"):
        print("\n[6/6] Importing 省控线汇总2022-2024 …")
        summary = import_control_2224(conn, data_dir)
        all_summary.update(summary)

    conn.close()

    # Final summary
    total = sum(all_summary.values())
    print("\n" + "=" * 60)
    print("IMPORT SUMMARY")
    print("=" * 60)
    for fname, cnt in all_summary.items():
        print(f"  {fname:<60s}  {cnt:>7,} records")
    print("-" * 60)
    print(f"  {'TOTAL':<60s}  {total:>7,} records")
    print("=" * 60)


if __name__ == "__main__":
    main()
